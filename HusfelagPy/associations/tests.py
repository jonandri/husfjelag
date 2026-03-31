from django.test import TestCase, Client
from unittest.mock import patch, MagicMock
from django.db import models as django_models
from .models import Association, HMSImportSource, Apartment, Category, BankAccount, Transaction, TransactionStatus, AssociationAccess, Budget, BudgetItem, ApartmentOwnership, Collection, CollectionStatus
from .scraper import scrape_hms_apartments
import json
import logging
from users.models import User


class HMSImportSourceModelTest(TestCase):
    def setUp(self):
        self.association = Association.objects.create(
            ssn="1234567890", name="Test Húsfélag",
            address="Testgata 1", postal_code="101", city="Reykjavík"
        )

    def test_create_source(self):
        src = HMSImportSource.objects.create(
            association=self.association,
            url="https://hms.is/fasteignaskra/228369/1203373",
            landeign_id=228369,
            stadfang_id=1203373,
        )
        self.assertEqual(src.association, self.association)
        self.assertEqual(src.landeign_id, 228369)
        self.assertEqual(src.stadfang_id, 1203373)
        self.assertIsNotNone(src.last_imported_at)

    def test_unique_together(self):
        HMSImportSource.objects.create(
            association=self.association,
            url="https://hms.is/fasteignaskra/228369/1203373",
            landeign_id=228369,
            stadfang_id=1203373,
        )
        from django.db import IntegrityError
        with self.assertRaises(IntegrityError):
            HMSImportSource.objects.create(
                association=self.association,
                url="https://hms.is/fasteignaskra/228369/1203373",
                landeign_id=228369,
                stadfang_id=1203373,
            )


class ScrapeHMSApartmentsTest(TestCase):

    def _make_api_response(self, fasteignir):
        """Minimal JSON mimicking the hms.is stadfang API response."""
        return {"stadfangData": {"fasteignir": fasteignir}}

    def test_scrape_returns_apartments(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = self._make_api_response([
            {"fasteign_nr": 2011134, "merking": "010101", "einflm": 68.5},
            {"fasteign_nr": 2011135, "merking": "020101", "einflm": 72.0},
        ])
        with patch("associations.scraper.requests.get", return_value=mock_resp):
            result = scrape_hms_apartments("https://hms.is/fasteignaskra/228369/1203373")
        self.assertIsNotNone(result)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["fnr"], "2011134")
        self.assertEqual(result[0]["anr"], "01 0101")
        self.assertAlmostEqual(float(result[0]["size"]), 68.5)

    def test_scrape_returns_none_on_error(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 500
        with patch("associations.scraper.requests.get", return_value=mock_resp):
            result = scrape_hms_apartments("https://hms.is/fasteignaskra/228369/1203373")
        self.assertIsNone(result)

    def test_scrape_returns_empty_list_when_no_rows(self):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = self._make_api_response([])
        with patch("associations.scraper.requests.get", return_value=mock_resp):
            result = scrape_hms_apartments("https://hms.is/fasteignaskra/228369/1203373")
        self.assertEqual(result, [])

    def test_scrape_returns_none_on_connection_error(self):
        import requests as req_lib
        with patch("associations.scraper.requests.get", side_effect=req_lib.RequestException):
            result = scrape_hms_apartments("https://hms.is/fasteignaskra/228369/1203373")
        self.assertIsNone(result)


class ImportPreviewViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="1234567890", name="Test User")
        self.association = Association.objects.create(
            ssn="0987654321", name="Test Húsfélag",
            address="Testgata 1", postal_code="101", city="Reykjavík"
        )
        from associations.models import AssociationAccess, AssociationRole
        AssociationAccess.objects.create(
            user=self.user, association=self.association,
            role=AssociationRole.CHAIR, active=True
        )

    def test_preview_classifies_create_update_missing(self):
        # existing apartment in DB
        Apartment.objects.create(
            association=self.association, fnr="2011135", anr="0201", size=70.0
        )
        scraped = [
            {"fnr": "2011134", "anr": "0101", "size": 68.5},  # new
            {"fnr": "2011135", "anr": "0201", "size": 72.0},  # update
        ]
        with patch("associations.views.scrape_hms_apartments", return_value=scraped):
            resp = self.client.post(
                "/Apartment/import/preview",
                data=json.dumps({"user_id": self.user.id, "urls": ["https://hms.is/fasteignaskra/228369/1203373"]}),
                content_type="application/json"
            )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data["create"]), 1)
        self.assertEqual(data["create"][0]["fnr"], "2011134")
        self.assertEqual(len(data["update"]), 1)
        self.assertEqual(data["update"][0]["fnr"], "2011135")
        self.assertEqual(len(data["missing"]), 0)

    def test_preview_reports_missing_apartment(self):
        Apartment.objects.create(
            association=self.association, fnr="2011099", anr="0301", size=55.0
        )
        with patch("associations.views.scrape_hms_apartments", return_value=[]):
            resp = self.client.post(
                "/Apartment/import/preview",
                data=json.dumps({"user_id": self.user.id, "urls": ["https://hms.is/fasteignaskra/228369/1203373"]}),
                content_type="application/json"
            )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data["missing"]), 1)
        self.assertEqual(data["missing"][0]["fnr"], "2011099")

    def test_preview_invalid_url_returns_400(self):
        resp = self.client.post(
            "/Apartment/import/preview",
            data=json.dumps({"user_id": self.user.id, "urls": ["https://example.com/bad"]}),
            content_type="application/json"
        )
        self.assertEqual(resp.status_code, 400)

    def test_preview_returns_502_when_hms_unreachable(self):
        # Suppress django.request logger to avoid Python 3.14/Django 4.1 debug-template crash on 5xx responses
        with patch("associations.views.scrape_hms_apartments", return_value=None):
            with self.assertLogs("django.request", level=logging.ERROR):
                resp = self.client.post(
                    "/Apartment/import/preview",
                    data=json.dumps({"user_id": self.user.id, "urls": ["https://hms.is/fasteignaskra/228369/1203373"]}),
                    content_type="application/json"
                )
        self.assertEqual(resp.status_code, 502)

    def test_preview_missing_urls_returns_400(self):
        resp = self.client.post(
            "/Apartment/import/preview",
            data=json.dumps({"user_id": self.user.id, "urls": []}),
            content_type="application/json"
        )
        self.assertEqual(resp.status_code, 400)

    def test_sources_returns_saved_urls(self):
        HMSImportSource.objects.create(
            association=self.association,
            url="https://hms.is/fasteignaskra/228369/1203373",
            landeign_id=228369,
            stadfang_id=1203373,
        )
        resp = self.client.get(f"/Apartment/import/sources?user_id={self.user.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["url"], "https://hms.is/fasteignaskra/228369/1203373")
        self.assertEqual(data[0]["landeign_id"], 228369)
        self.assertEqual(data[0]["stadfang_id"], 1203373)


class ImportConfirmViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="1111111111", name="Confirm User")
        self.association = Association.objects.create(
            ssn="2222222222", name="Confirm Húsfélag",
            address="Confirmgata 2", postal_code="200", city="Kópavogur"
        )
        from associations.models import AssociationAccess, AssociationRole
        AssociationAccess.objects.create(
            user=self.user, association=self.association,
            role=AssociationRole.CHAIR, active=True
        )

    def test_confirm_creates_new_apartments(self):
        scraped = [{"fnr": "3011100", "anr": "0101", "size": 50.0}]
        with patch("associations.views.scrape_hms_apartments", return_value=scraped):
            resp = self.client.post(
                "/Apartment/import/confirm",
                data=json.dumps({
                    "user_id": self.user.id,
                    "urls": ["https://hms.is/fasteignaskra/100/200"],
                    "deactivate_ids": []
                }),
                content_type="application/json"
            )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(
            self.association.apartments.filter(fnr="3011100", deleted=False).exists()
        )

    def test_confirm_updates_existing_apartment(self):
        apt = Apartment.objects.create(
            association=self.association, fnr="3011101", anr="OLD", size=40.0
        )
        scraped = [{"fnr": "3011101", "anr": "NEW", "size": 45.0}]
        with patch("associations.views.scrape_hms_apartments", return_value=scraped):
            resp = self.client.post(
                "/Apartment/import/confirm",
                data=json.dumps({
                    "user_id": self.user.id,
                    "urls": ["https://hms.is/fasteignaskra/100/200"],
                    "deactivate_ids": []
                }),
                content_type="application/json"
            )
        self.assertEqual(resp.status_code, 200)
        apt.refresh_from_db()
        self.assertEqual(apt.anr, "NEW")
        self.assertAlmostEqual(float(apt.size), 45.0)

    def test_confirm_deactivates_selected_apartments(self):
        apt = Apartment.objects.create(
            association=self.association, fnr="3011199", anr="0901", size=60.0
        )
        with patch("associations.views.scrape_hms_apartments", return_value=[]):
            resp = self.client.post(
                "/Apartment/import/confirm",
                data=json.dumps({
                    "user_id": self.user.id,
                    "urls": ["https://hms.is/fasteignaskra/100/200"],
                    "deactivate_ids": [apt.id]
                }),
                content_type="application/json"
            )
        self.assertEqual(resp.status_code, 200)
        apt.refresh_from_db()
        self.assertTrue(apt.deleted)

    def test_confirm_saves_hms_source(self):
        with patch("associations.views.scrape_hms_apartments", return_value=[]):
            self.client.post(
                "/Apartment/import/confirm",
                data=json.dumps({
                    "user_id": self.user.id,
                    "urls": ["https://hms.is/fasteignaskra/100/200"],
                    "deactivate_ids": []
                }),
                content_type="application/json"
            )
        src = HMSImportSource.objects.get(association=self.association, stadfang_id=200)
        self.assertEqual(src.landeign_id, 100)
        self.assertEqual(src.url, "https://hms.is/fasteignaskra/100/200")


class CategoryGlobalModelTest(TestCase):
    def test_category_has_no_association_field(self):
        """Category can be created without an association."""
        from associations.models import Category
        cat = Category.objects.create(name="Tryggingar", type="SHARED")
        self.assertEqual(cat.name, "Tryggingar")
        self.assertFalse(hasattr(cat, 'association_id'))

    def test_two_categories_same_name_allowed(self):
        """Without unique_together, duplicate names across old associations are fine."""
        from associations.models import Category
        Category.objects.create(name="Hiti", type="SHARE2")
        Category.objects.create(name="Hiti", type="SHARE2")  # should not raise
        self.assertEqual(Category.objects.filter(name="Hiti").count(), 2)


class CategoryListViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        from associations.models import Category
        Category.objects.create(name="Tryggingar", type="SHARED")
        Category.objects.create(name="Hiti", type="SHARE2")
        Category.objects.create(name="Óvirkur", type="EQUAL", deleted=True)

    def test_list_returns_only_active_categories(self):
        resp = self.client.get("/Category/list")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data), 2)
        names = {c["name"] for c in data}
        self.assertIn("Tryggingar", names)
        self.assertIn("Hiti", names)
        self.assertNotIn("Óvirkur", names)

    def test_list_returns_id_name_type(self):
        resp = self.client.get("/Category/list")
        item = resp.json()[0]
        self.assertIn("id", item)
        self.assertIn("name", item)
        self.assertIn("type", item)


class CategorySuperadminGuardTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.superadmin = User.objects.create(kennitala="0000000001", name="Super", is_superadmin=True)
        self.regular = User.objects.create(kennitala="0000000002", name="Regular", is_superadmin=False)

    def test_post_category_requires_superadmin(self):
        resp = self.client.post(
            "/Category",
            data=json.dumps({"user_id": self.regular.id, "name": "Test", "type": "SHARED"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_post_category_succeeds_for_superadmin(self):
        resp = self.client.post(
            "/Category",
            data=json.dumps({"user_id": self.superadmin.id, "name": "Tryggingar", "type": "SHARED"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()["name"], "Tryggingar")

    def test_put_category_requires_superadmin(self):
        from associations.models import Category
        cat = Category.objects.create(name="Old", type="SHARED")
        resp = self.client.put(
            f"/Category/update/{cat.id}?user_id={self.regular.id}",
            data=json.dumps({"name": "New", "type": "SHARED"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_delete_category_requires_superadmin(self):
        from associations.models import Category
        cat = Category.objects.create(name="ToDelete", type="EQUAL")
        resp = self.client.delete(f"/Category/delete/{cat.id}?user_id={self.regular.id}")
        self.assertEqual(resp.status_code, 403)
        cat.refresh_from_db()
        self.assertFalse(cat.deleted)

    def test_enable_category_requires_superadmin(self):
        from associations.models import Category
        cat = Category.objects.create(name="Disabled", type="EQUAL", deleted=True)
        resp = self.client.patch(f"/Category/enable/{cat.id}?user_id={self.regular.id}")
        self.assertEqual(resp.status_code, 403)

    def test_post_category_missing_user_id_returns_400(self):
        resp = self.client.post(
            "/Category",
            data=json.dumps({"name": "Test", "type": "SHARED"}),  # no user_id
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)

    def test_put_category_superadmin_succeeds(self):
        from associations.models import Category
        cat = Category.objects.create(name="Old", type="SHARED")
        resp = self.client.put(
            f"/Category/update/{cat.id}?user_id={self.superadmin.id}",
            data=json.dumps({"name": "New", "type": "SHARE2"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["name"], "New")

    def test_delete_category_superadmin_succeeds(self):
        from associations.models import Category
        cat = Category.objects.create(name="ToDelete", type="EQUAL")
        resp = self.client.delete(f"/Category/delete/{cat.id}?user_id={self.superadmin.id}")
        self.assertEqual(resp.status_code, 204)
        cat.refresh_from_db()
        self.assertTrue(cat.deleted)

    def test_enable_category_superadmin_succeeds(self):
        from associations.models import Category
        cat = Category.objects.create(name="Disabled", type="SHARED", deleted=True)
        resp = self.client.patch(f"/Category/enable/{cat.id}?user_id={self.superadmin.id}")
        self.assertEqual(resp.status_code, 200)
        cat.refresh_from_db()
        self.assertFalse(cat.deleted)


class BudgetWizardViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="9999999901", name="Wizard User")
        self.association = Association.objects.create(
            ssn="9999999902", name="Wizard Húsfélag",
            address="Wizardgata 1", postal_code="600", city="Akureyri"
        )
        from associations.models import AssociationAccess, AssociationRole, Category
        AssociationAccess.objects.create(
            user=self.user, association=self.association,
            role=AssociationRole.CHAIR, active=True
        )
        self.cat1 = Category.objects.create(name="Tryggingar", type="SHARED")
        self.cat2 = Category.objects.create(name="Hiti", type="SHARE2")

    def test_wizard_creates_budget_with_items(self):
        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": self.user.id,
                "items": [
                    {"category_id": self.cat1.id, "amount": 450000},
                    {"category_id": self.cat2.id, "amount": 120000},
                ]
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(data["is_active"], True)
        self.assertEqual(data["version"], 1)
        self.assertEqual(len(data["items"]), 2)
        amounts = {i["category_id"]: float(i["amount"]) for i in data["items"]}
        self.assertAlmostEqual(amounts[self.cat1.id], 450000)
        self.assertAlmostEqual(amounts[self.cat2.id], 120000)

    def test_wizard_deactivates_previous_budget(self):
        from associations.models import Budget
        old = Budget.objects.create(
            association=self.association, year=2025, version=1, is_active=True
        )
        import datetime
        year = datetime.date.today().year
        old.year = year
        old.save()

        self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": self.user.id,
                "items": [{"category_id": self.cat1.id, "amount": 100}],
            }),
            content_type="application/json",
        )
        old.refresh_from_db()
        self.assertFalse(old.is_active)

    def test_wizard_increments_version(self):
        from associations.models import Budget
        import datetime
        year = datetime.date.today().year
        Budget.objects.create(association=self.association, year=year, version=1, is_active=True)

        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": self.user.id,
                "items": [{"category_id": self.cat1.id, "amount": 1}],
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()["version"], 2)

    def test_wizard_returns_400_for_empty_items(self):
        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({"user_id": self.user.id, "items": []}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)

    def test_wizard_returns_400_for_invalid_category(self):
        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": self.user.id,
                "items": [{"category_id": 99999, "amount": 100}],
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)

    def test_wizard_returns_404_for_unknown_user(self):
        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": 99999,
                "items": [{"category_id": self.cat1.id, "amount": 100}],
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 404)

    def test_budget_get_returns_null_when_no_budget(self):
        resp = self.client.get(f"/Budget/{self.user.id}")
        self.assertEqual(resp.status_code, 200)
        # DRF renders Response(None) as empty body; assert no budget data returned
        self.assertFalse(resp.content)

    def test_wizard_returns_400_for_negative_amount(self):
        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": self.user.id,
                "items": [{"category_id": self.cat1.id, "amount": -1}],
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)

    def test_wizard_returns_400_for_non_numeric_amount(self):
        resp = self.client.post(
            "/Budget/wizard",
            data=json.dumps({
                "user_id": self.user.id,
                "items": [{"category_id": self.cat1.id, "amount": "bad"}],
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)


class AccountingKeyModelTest(TestCase):
    def test_create_accounting_key(self):
        from associations.models import AccountingKey, AccountingKeyType
        key = AccountingKey.objects.create(
            number=9990, name="Test lykill", type=AccountingKeyType.EXPENSE
        )
        self.assertEqual(key.number, 9990)
        self.assertEqual(key.name, "Test lykill")
        self.assertEqual(key.type, "EXPENSE")
        self.assertFalse(key.deleted)

    def test_seed_data_present(self):
        from associations.models import AccountingKey
        # Migration seeds 12 keys; verify a few
        self.assertTrue(AccountingKey.objects.filter(number=1200).exists())
        self.assertTrue(AccountingKey.objects.filter(number=5600).exists())
        self.assertEqual(AccountingKey.objects.filter(deleted=False).count(), 12)

    def test_ordering_by_number(self):
        from associations.models import AccountingKey
        # Seeded keys should come back ordered by number
        keys = list(AccountingKey.objects.all().values_list("number", flat=True))
        self.assertEqual(keys, sorted(keys))


class AccountingKeyViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.superadmin = User.objects.create(
            kennitala="1111111111", name="Admin", is_superadmin=True
        )
        self.regular = User.objects.create(
            kennitala="2222222222", name="Regular"
        )
        # Use numbers outside the seeded range to avoid collisions
        from associations.models import AccountingKey, AccountingKeyType
        self.key = AccountingKey.objects.create(
            number=9901, name="Test Eign", type=AccountingKeyType.ASSET
        )
        self.deleted_key = AccountingKey.objects.create(
            number=9902, name="Test Óvirkur", type=AccountingKeyType.EXPENSE, deleted=True
        )

    def test_list_returns_only_active_keys(self):
        resp = self.client.get("/AccountingKey/list")
        self.assertEqual(resp.status_code, 200)
        numbers = [k["number"] for k in resp.json()]
        self.assertIn(9901, numbers)
        self.assertNotIn(9902, numbers)

    def test_superadmin_get_includes_deleted(self):
        resp = self.client.get(f"/AccountingKey/{self.superadmin.id}")
        self.assertEqual(resp.status_code, 200)
        numbers = [k["number"] for k in resp.json()]
        self.assertIn(9901, numbers)
        self.assertIn(9902, numbers)

    def test_create_requires_superadmin(self):
        resp = self.client.post(
            "/AccountingKey",
            data=json.dumps({"user_id": self.regular.id, "number": 9999, "name": "X", "type": "EXPENSE"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_create_accounting_key(self):
        resp = self.client.post(
            "/AccountingKey",
            data=json.dumps({"user_id": self.superadmin.id, "number": 9950, "name": "Nýr lykill", "type": "EXPENSE"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()["number"], 9950)
        self.assertEqual(resp.json()["type"], "EXPENSE")

    def test_create_duplicate_number_returns_400(self):
        resp = self.client.post(
            "/AccountingKey",
            data=json.dumps({"user_id": self.superadmin.id, "number": 9901, "name": "Afrit", "type": "ASSET"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("þegar til", resp.json()["detail"])

    def test_update_accounting_key(self):
        resp = self.client.put(
            f"/AccountingKey/update/{self.key.id}?user_id={self.superadmin.id}",
            data=json.dumps({"name": "Uppfært nafn", "type": "ASSET"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["name"], "Uppfært nafn")

    def test_soft_delete(self):
        resp = self.client.delete(
            f"/AccountingKey/delete/{self.key.id}?user_id={self.superadmin.id}"
        )
        self.assertEqual(resp.status_code, 204)
        self.key.refresh_from_db()
        self.assertTrue(self.key.deleted)

    def test_enable(self):
        resp = self.client.patch(
            f"/AccountingKey/enable/{self.deleted_key.id}?user_id={self.superadmin.id}"
        )
        self.assertEqual(resp.status_code, 200)
        self.deleted_key.refresh_from_db()
        self.assertFalse(self.deleted_key.deleted)

    def test_non_superadmin_cannot_delete(self):
        resp = self.client.delete(
            f"/AccountingKey/delete/{self.key.id}?user_id={self.regular.id}"
        )
        self.assertEqual(resp.status_code, 403)


class CategoryAccountingFKTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.superadmin = User.objects.create(
            kennitala="3333333333", name="Admin", is_superadmin=True
        )
        from associations.models import AccountingKey, AccountingKeyType, Category, CategoryType
        self.expense_key = AccountingKey.objects.create(
            number=9801, name="Test Gjöld", type=AccountingKeyType.EXPENSE
        )
        self.income_key = AccountingKey.objects.create(
            number=9802, name="Test Tekjur", type=AccountingKeyType.INCOME
        )
        self.category = Category.objects.create(name="Þrif", type=CategoryType.SHARED)

    def test_category_has_accounting_fks(self):
        from associations.models import Category
        cat = Category.objects.get(id=self.category.id)
        self.assertIsNone(cat.expense_account)
        self.assertIsNone(cat.income_account)

    def test_update_category_sets_expense_account(self):
        resp = self.client.put(
            f"/Category/update/{self.category.id}?user_id={self.superadmin.id}",
            data=json.dumps({
                "name": "Þrif",
                "type": "SHARED",
                "expense_account_id": self.expense_key.id,
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["expense_account_id"], self.expense_key.id)
        self.assertEqual(data["expense_account_number"], 9801)

    def test_update_category_clears_expense_account(self):
        self.category.expense_account = self.expense_key
        self.category.save()
        resp = self.client.put(
            f"/Category/update/{self.category.id}?user_id={self.superadmin.id}",
            data=json.dumps({"name": "Þrif", "type": "SHARED", "expense_account_id": None}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIsNone(resp.json()["expense_account_id"])

    def test_serializer_returns_account_fields(self):
        self.category.expense_account = self.expense_key
        self.category.income_account = self.income_key
        self.category.save()
        resp = self.client.get(f"/Category/{self.superadmin.id}")
        self.assertEqual(resp.status_code, 200)
        cat = next(c for c in resp.json() if c["id"] == self.category.id)
        self.assertEqual(cat["expense_account_id"], self.expense_key.id)
        self.assertEqual(cat["income_account_id"], self.income_key.id)


class BankAccountViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="4444444444", name="Formaður")
        self.other_user = User.objects.create(kennitala="5555555555", name="Annar")
        self.association = Association.objects.create(
            ssn="1111111119", name="Test HF", address="Testgata 1",
            postal_code="101", city="Reykjavík"
        )
        self.other_association = Association.objects.create(
            ssn="2222222228", name="Annað HF", address="Testgata 2",
            postal_code="101", city="Reykjavík"
        )
        from associations.models import AssociationAccess, AssociationRole, AccountingKey, AccountingKeyType
        AssociationAccess.objects.create(
            user=self.user, association=self.association,
            role=AssociationRole.CHAIR, active=True
        )
        AssociationAccess.objects.create(
            user=self.other_user, association=self.other_association,
            role=AssociationRole.CHAIR, active=True
        )
        self.asset_key = AccountingKey.objects.create(
            number=9701, name="Test Reikningur", type=AccountingKeyType.ASSET
        )

    def test_create_bank_account(self):
        resp = self.client.post(
            "/BankAccount",
            data=json.dumps({
                "user_id": self.user.id,
                "name": "Rekstrarreikningur",
                "account_number": "0101-26-123456",
                "asset_account_id": self.asset_key.id,
                "description": "Aðalreikningur",
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(data["name"], "Rekstrarreikningur")
        self.assertEqual(data["asset_account"]["number"], 9701)

    def test_list_bank_accounts(self):
        from associations.models import BankAccount
        BankAccount.objects.create(
            association=self.association, name="Rekstrar",
            account_number="0101-26-123456", asset_account=self.asset_key
        )
        resp = self.client.get(f"/BankAccount/{self.user.id}?as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.json()), 1)

    def test_list_excludes_deleted(self):
        from associations.models import BankAccount
        BankAccount.objects.create(
            association=self.association, name="Gamall", account_number="0000-00-000000", deleted=True
        )
        resp = self.client.get(f"/BankAccount/{self.user.id}?as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.json()), 0)

    def test_update_bank_account(self):
        from associations.models import BankAccount
        bank = BankAccount.objects.create(
            association=self.association, name="Gamalt nafn", account_number="0101-26-123456"
        )
        resp = self.client.put(
            f"/BankAccount/update/{bank.id}",
            data=json.dumps({
                "user_id": self.user.id,
                "name": "Nýtt nafn",
                "account_number": "0101-26-999999",
                "asset_account_id": self.asset_key.id,
                "description": "",
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["name"], "Nýtt nafn")

    def test_delete_bank_account(self):
        from associations.models import BankAccount
        bank = BankAccount.objects.create(
            association=self.association, name="Rekstrar", account_number="0101-26-123456"
        )
        resp = self.client.delete(
            f"/BankAccount/delete/{bank.id}",
            data=json.dumps({"user_id": self.user.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 204)
        bank.refresh_from_db()
        self.assertTrue(bank.deleted)

    def test_cannot_delete_other_associations_bank_account(self):
        from associations.models import BankAccount
        bank = BankAccount.objects.create(
            association=self.other_association, name="Rekstrar", account_number="0101-26-999999"
        )
        resp = self.client.delete(
            f"/BankAccount/delete/{bank.id}",
            data=json.dumps({"user_id": self.user.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_no_association_returns_empty_list(self):
        nobody = User.objects.create(kennitala="6666666666", name="Nobody")
        resp = self.client.get(f"/BankAccount/{nobody.id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), [])


class TransactionViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="7777777777", name="Gjaldkeri")
        self.association = Association.objects.create(
            ssn="3333333337", name="Felag HF", address="Gata 1",
            postal_code="101", city="Reykjavík"
        )
        from associations.models import AssociationAccess, AssociationRole, AccountingKey, AccountingKeyType, BankAccount, Category, CategoryType
        AssociationAccess.objects.create(
            user=self.user, association=self.association,
            role=AssociationRole.CFO, active=True
        )
        self.bank_account = BankAccount.objects.create(
            association=self.association,
            name="Rekstrar",
            account_number="0101-26-123456",
        )
        self.category = Category.objects.create(name="Tryggingar", type=CategoryType.SHARED)

    def test_create_manual_transaction(self):
        resp = self.client.post(
            "/Transaction",
            data=json.dumps({
                "user_id": self.user.id,
                "bank_account_id": self.bank_account.id,
                "date": "2026-03-15",
                "amount": "-180000.00",
                "description": "VÍS tryggingar",
                "reference": "REF001",
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(data["description"], "VÍS tryggingar")
        self.assertEqual(data["status"], "IMPORTED")
        self.assertEqual(data["bank_account"]["name"], "Rekstrar")

    def test_create_with_category_sets_categorised_status(self):
        resp = self.client.post(
            "/Transaction",
            data=json.dumps({
                "user_id": self.user.id,
                "bank_account_id": self.bank_account.id,
                "date": "2026-03-15",
                "amount": "-50000.00",
                "description": "Þrif",
                "category_id": self.category.id,
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()["status"], "CATEGORISED")

    def test_list_transactions(self):
        from associations.models import Transaction
        Transaction.objects.create(
            bank_account=self.bank_account,
            date="2026-03-01",
            amount="-10000",
            description="Test",
            status="IMPORTED",
        )
        resp = self.client.get(f"/Transaction/{self.user.id}?as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.json()), 1)

    def test_list_filters_by_year(self):
        from associations.models import Transaction
        Transaction.objects.create(
            bank_account=self.bank_account, date="2025-06-01",
            amount="-1000", description="Gamla", status="IMPORTED"
        )
        Transaction.objects.create(
            bank_account=self.bank_account, date="2026-01-01",
            amount="-2000", description="Nýja", status="IMPORTED"
        )
        resp = self.client.get(f"/Transaction/{self.user.id}?as={self.association.id}&year=2026")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["description"], "Nýja")

    def test_categorise_transaction(self):
        from associations.models import Transaction
        tx = Transaction.objects.create(
            bank_account=self.bank_account, date="2026-03-01",
            amount="-5000", description="Test", status="IMPORTED"
        )
        resp = self.client.patch(
            f"/Transaction/categorise/{tx.id}",
            data=json.dumps({"user_id": self.user.id, "category_id": self.category.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["status"], "CATEGORISED")
        self.assertEqual(data["category"]["id"], self.category.id)

    def test_categorise_wrong_category_returns_404(self):
        from associations.models import Transaction
        tx = Transaction.objects.create(
            bank_account=self.bank_account, date="2026-03-01",
            amount="-5000", description="Test", status="IMPORTED"
        )
        resp = self.client.patch(
            f"/Transaction/categorise/{tx.id}",
            data=json.dumps({"user_id": self.user.id, "category_id": 99999}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 404)

    def test_no_bank_accounts_returns_empty_list(self):
        nobody = User.objects.create(kennitala="8888888888", name="Nobody")
        nobody_assoc = Association.objects.create(
            ssn="4444444446", name="Empty HF", address="Gata 1",
            postal_code="101", city="Reykjavík"
        )
        from associations.models import AssociationAccess, AssociationRole
        AssociationAccess.objects.create(
            user=nobody, association=nobody_assoc,
            role=AssociationRole.CHAIR, active=True
        )
        resp = self.client.get(f"/Transaction/{nobody.id}?as={nobody_assoc.id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), [])


class ImporterTest(TestCase):
    """Unit tests for importers.py — no HTTP, no DB (except detect_duplicates)."""

    def test_parse_icelandic_amount_comma_decimal(self):
        from associations.importers import parse_icelandic_amount
        from decimal import Decimal
        self.assertEqual(parse_icelandic_amount("-100,00"), Decimal("-100.00"))
        self.assertEqual(parse_icelandic_amount("455,00"), Decimal("455.00"))
        self.assertEqual(parse_icelandic_amount("-351.427,00"), Decimal("-351427.00"))

    def test_parse_icelandic_amount_kr_suffix(self):
        from associations.importers import parse_icelandic_amount
        from decimal import Decimal
        self.assertEqual(parse_icelandic_amount("-300 kr."), Decimal("-300"))
        self.assertEqual(parse_icelandic_amount("-2.805.615 kr."), Decimal("-2805615"))
        self.assertEqual(parse_icelandic_amount("1.135.983 kr."), Decimal("1135983"))

    def test_parse_icelandic_amount_float(self):
        from associations.importers import parse_icelandic_amount
        from decimal import Decimal
        self.assertEqual(parse_icelandic_amount(-100.0), Decimal("-100"))
        self.assertEqual(parse_icelandic_amount(245000.0), Decimal("245000"))

    def test_parse_arion_csv(self):
        from associations.importers import parse_arion
        from decimal import Decimal
        import datetime
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_bytes = (
            ";;\n"
            "0370-13-037063;IS87 0370 1303 7063 0507 7253 59\n"
            ";;\n"
            "Dagsetning;Upphæð;Staða;Mynt;Skýring;Seðilnúmer;Tilvísun;Texti\n"
            "15.03.2026;-245.000,00;;;HS Veitur hf.;280226;;HS Veitur hf.\n"
            "10.03.2026;320.000,00;;;Innborgun;310326;;Innborgun\n"
        ).encode("utf-8")
        f = SimpleUploadedFile("AccountTransactions0370.csv", csv_bytes, content_type="text/csv")
        result = parse_arion(f, "csv")
        self.assertEqual(result["file_account_number"], "0370-13-037063")
        self.assertEqual(len(result["rows"]), 2)
        self.assertEqual(result["rows"][0]["date"], datetime.date(2026, 3, 15))
        self.assertEqual(result["rows"][0]["amount"], Decimal("-245000.00"))
        self.assertEqual(result["rows"][0]["description"], "HS Veitur hf.")
        self.assertEqual(result["rows"][0]["reference"], "280226")
        self.assertEqual(result["rows"][1]["amount"], Decimal("320000.00"))

    def test_parse_arion_xlsx(self):
        from associations.importers import parse_arion
        from decimal import Decimal
        import datetime
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Heiti", "IBAN númer"])
        ws.append(["0370-13-037063", "IS87 0370 1303 7063 0507 7253 59"])
        ws.append([None])
        ws.append(["Dagsetning", "Upphæð", "Staða", "Mynt", "Skýring", "Seðilnúmer"])
        ws.append([datetime.datetime(2026, 3, 15), -245000.0, None, "ISK", "HS Veitur hf.", "280226"])
        buf = BytesIO()
        wb.save(buf)
        f = SimpleUploadedFile("AccountTransactions0370.xlsx", buf.getvalue(),
                               content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        result = parse_arion(f, "xlsx")
        self.assertEqual(result["file_account_number"], "0370-13-037063")
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["date"], datetime.date(2026, 3, 15))
        self.assertEqual(result["rows"][0]["amount"], Decimal("-245000"))
        self.assertEqual(result["rows"][0]["description"], "HS Veitur hf.")

    def test_parse_landsbankinn_csv(self):
        from associations.importers import parse_landsbankinn
        from decimal import Decimal
        import datetime
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_bytes = (
            "Netbanki fyrirtækja-Reikningsyfirlit\n"
            "Færslur á reikningi 0133-26-019111 Veltureikningur fyrirtækja\n"
            "Allar færslur. Tímabil 28.3.2025 - 29.3.2026\n"
            "\n"
            "Dags;Vaxtad;Banki;RB. Nr.;Fl.;Tnr/Seðilnr.;Tilvísun;Textalykill;Skýring greiðslu;Kennitala;Texti;Upphæð;Staða\n"
            "24.03.2026;24.03;0536;^h71;01;0010426;2405862319;Félagaþjónusta;Félagaþjónusta;240586-2319;Hilmar Þór Birgisson;24.484;242.562\n"
            "23.03.2026;23.03;0133;KR41;02;0002012;2312080590;Rafmagn og hiti;Rafmagn og hiti;431208-0590;HS Veitur hf.;-1.948;218.078\n"
        ).encode("utf-8")
        f = SimpleUploadedFile("LandsbankinnExcel20260330.csv", csv_bytes)
        result = parse_landsbankinn(f, "csv")
        self.assertEqual(result["file_account_number"], "0133-26-019111")
        self.assertEqual(len(result["rows"]), 2)
        self.assertEqual(result["rows"][0]["date"], datetime.date(2026, 3, 24))
        self.assertEqual(result["rows"][0]["description"], "Hilmar Þór Birgisson")
        self.assertEqual(result["rows"][0]["reference"], "0010426")
        self.assertEqual(result["rows"][1]["amount"], Decimal("-1948"))
        # When Texti is empty, falls back to Skýring greiðslu
        csv_no_texti = (
            "Netbanki\n"
            "Færslur á reikningi 0133-26-019111 Reikningur\n"
            "\n"
            "\n"
            "Dags;Vaxtad;Banki;RB. Nr.;Fl.;Tnr/Seðilnr.;Tilvísun;Textalykill;Skýring greiðslu;Kennitala;Texti;Upphæð;Staða\n"
            "01.03.2026;;;;;\t;;Kostnaður;HS Veitur hf.;;;-500;\n"
        ).encode("utf-8")
        f2 = SimpleUploadedFile("LandsbankinnExcel20260301.csv", csv_no_texti)
        result2 = parse_landsbankinn(f2, "csv")
        self.assertEqual(result2["rows"][0]["description"], "HS Veitur hf.")

    def test_parse_islandsbanki_new_format(self):
        from associations.importers import parse_islandsbanki
        from decimal import Decimal
        import datetime, openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Eigandi", "Þórunnarstræti 132, húsfélag"])  # row 1
        ws.append(["Kennitala", "650585-1279"])                   # row 2
        ws.append(["Reikningur", "Húsfélagar. Aðalreik"])         # row 3
        ws.append(["Reikningsnúmer", "0565-26-565121"])            # row 4 — detection key
        ws.append(["Staða", "189.153 kr."])                        # row 5
        ws.append([None])                                           # row 6
        ws.append(["Dagsetning frá", "28.02.2026"])                # row 7
        ws.append(["Dagsetning til", "30.03.2026"])                # row 8
        ws.append(["Yfirlit sótt", "2026-03-30 09:28:35"])         # row 9
        ws.append([None])                                           # row 10
        ws.append([None])                                           # row 11
        ws.append(["Dagsetning", "Mótaðili", "Tilvísun", "Texti", "Upphæð", "Staða"])  # row 12
        ws.append(["18.03.2026", "LukTom píparar ehf.", "280226", "Kostnaður", "-300 kr.", "189.153 kr."])  # row 13
        ws.append(["17.03.2026", "Þjónustugjald innheimtuþjónusta", None, "Innheimtuþjónusta", "-919 kr.", "189.453 kr."])  # row 14
        buf = BytesIO()
        wb.save(buf)
        f = SimpleUploadedFile("reikningsyfirlit2026-03-30.xlsx", buf.getvalue())
        result = parse_islandsbanki(f, "xlsx")
        self.assertEqual(result["file_account_number"], "0565-26-565121")
        self.assertEqual(len(result["rows"]), 2)
        self.assertEqual(result["rows"][0]["date"], datetime.date(2026, 3, 18))
        self.assertEqual(result["rows"][0]["description"], "LukTom píparar ehf.")
        self.assertEqual(result["rows"][0]["amount"], Decimal("-300"))
        self.assertEqual(result["rows"][0]["reference"], "280226")

    def test_parse_islandsbanki_old_format(self):
        from associations.importers import parse_islandsbanki
        from decimal import Decimal
        import datetime, openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Íslandsbanki"])                           # row 1
        ws.append(["Reikningsyfirlit"])                        # row 2
        ws.append(["Tímabil: mars 2026"])                      # row 3
        ws.append(["Allar færslur"])                           # row 4 — NOT "Reikningsnúmer"
        ws.append(["Dags.", "Seðilnr.", "Tegund", "Mótaðili", "Tilvísun",
                   "Upplýsingar um færslu", "Aðrar upplýsingar", "Færslulykill",
                   "Textalykill", "Upplýsingar", "Kennitala móttakanda", "Söluaðili",
                   "Innlausnarbanki", "Vaxtadagsetning", "Bókunardagur greiðslu",
                   "Upphæð", "Upph.ISK", "Staða"])             # row 5 — headers
        ws.append(["18.03.2026", "280226", "Kostnaður", "LukTom píparar ehf.", "5603061130",
                   None, None, None,
                   "Kostnaður", "Innheimtukrafa", "6812221110", None,
                   None, None, None,
                   "-300", "-300", "189.153"])                 # row 6
        ws.append(["17.03.2026", "030326", "Millifært", "Elva Sturludóttir", "1607735109",
                   None, None, None,
                   "Millifært", "Innborgun", "1607735109", None,
                   None, None, None,
                   "1.135.983", "1.135.983", "5.785.724"])     # row 7
        buf = BytesIO()
        wb.save(buf)
        f = SimpleUploadedFile("ReikningsYfirlit20260330.xlsx", buf.getvalue())
        result = parse_islandsbanki(f, "xlsx")
        self.assertIsNone(result["file_account_number"])
        self.assertEqual(len(result["rows"]), 2)
        self.assertEqual(result["rows"][0]["date"], datetime.date(2026, 3, 18))
        self.assertEqual(result["rows"][0]["description"], "LukTom píparar ehf.")
        self.assertEqual(result["rows"][0]["amount"], Decimal("-300"))
        self.assertEqual(result["rows"][1]["amount"], Decimal("1135983"))

    def test_detect_duplicates(self):
        from associations.importers import detect_duplicates
        from associations.models import (
            Association, AssociationAccess, AssociationRole,
            AccountingKey, AccountingKeyType, BankAccount, Transaction
        )
        import datetime
        from decimal import Decimal
        assoc = Association.objects.create(
            ssn="9900000009", name="Dup Test HF", address="Gata 1",
            postal_code="101", city="Reykjavík"
        )
        asset_key = AccountingKey.objects.create(
            number=9750, name="Test", type=AccountingKeyType.ASSET
        )
        bank_account = BankAccount.objects.create(
            association=assoc, name="Test", account_number="0133-26-000001",
            asset_account=asset_key
        )
        Transaction.objects.create(
            bank_account=bank_account, date=datetime.date(2026, 3, 15),
            amount=Decimal("-245000.00"), description="HS Veitur hf.", status="IMPORTED"
        )
        rows = [
            {"date": datetime.date(2026, 3, 15), "amount": Decimal("-245000.00"),
             "description": "HS Veitur hf.", "reference": "280226"},   # duplicate
            {"date": datetime.date(2026, 3, 10), "amount": Decimal("-180000.00"),
             "description": "VÍS tryggingar", "reference": "290226"},  # new
        ]
        to_import, skipped = detect_duplicates(rows, bank_account)
        self.assertEqual(skipped, 1)
        self.assertEqual(len(to_import), 1)


class ImportViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="6600000001", name="Formaður")
        self.other_user = User.objects.create(kennitala="6600000002", name="Annar")
        self.association = Association.objects.create(
            ssn="6600000009", name="Import Test HF", address="Gata 1",
            postal_code="101", city="Reykjavík"
        )
        self.other_association = Association.objects.create(
            ssn="7700000009", name="Önnur HF", address="Gata 2",
            postal_code="101", city="Reykjavík"
        )
        from associations.models import AssociationAccess, AssociationRole, AccountingKey, AccountingKeyType
        AssociationAccess.objects.create(
            user=self.user, association=self.association,
            role=AssociationRole.CHAIR, active=True
        )
        AssociationAccess.objects.create(
            user=self.other_user, association=self.other_association,
            role=AssociationRole.CHAIR, active=True
        )
        self.asset_key = AccountingKey.objects.create(
            number=9760, name="Test Reikningur", type=AccountingKeyType.ASSET
        )
        from associations.models import BankAccount
        self.bank_account = BankAccount.objects.create(
            association=self.association,
            name="Rekstrarreikningur",
            account_number="0370-13-037063",
            asset_account=self.asset_key,
        )

    def _arion_csv(self, account_number="0370-13-037063", rows=None):
        """Build a minimal Arion CSV file as bytes."""
        from django.core.files.uploadedfile import SimpleUploadedFile
        if rows is None:
            rows = [
                "15.03.2026;-245.000,00;;;HS Veitur hf.;280226;;HS Veitur hf.",
                "10.03.2026;-180.000,00;;;VÍS tryggingar;290226;;VÍS tryggingar",
            ]
        lines = [
            "Heiti;IBAN",
            f"{account_number};IS87...",
            ";;",
            "Dagsetning;Upphæð;Staða;Mynt;Skýring;Seðilnúmer;Tilvísun;Texti",
        ] + rows
        return SimpleUploadedFile(
            "AccountTransactions0370.csv",
            "\n".join(lines).encode("utf-8"),
            content_type="text/csv",
        )

    def test_preview_returns_correct_counts(self):
        f = self._arion_csv()
        resp = self.client.post("/Import/preview", data={
            "user_id": self.user.id,
            "bank_account_id": self.bank_account.id,
            "bank": "arion",
            "file": f,
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["total_in_file"], 2)
        self.assertEqual(data["to_import"], 2)
        self.assertEqual(data["skipped_duplicates"], 0)
        self.assertEqual(len(data["rows"]), 2)
        self.assertEqual(data["rows"][0]["description"], "HS Veitur hf.")

    def test_preview_skips_duplicates(self):
        from associations.models import Transaction
        import datetime
        from decimal import Decimal
        Transaction.objects.create(
            bank_account=self.bank_account, date=datetime.date(2026, 3, 15),
            amount=Decimal("-245000.00"), description="HS Veitur hf.", status="IMPORTED"
        )
        f = self._arion_csv()
        resp = self.client.post("/Import/preview", data={
            "user_id": self.user.id,
            "bank_account_id": self.bank_account.id,
            "bank": "arion",
            "file": f,
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["total_in_file"], 2)
        self.assertEqual(data["to_import"], 1)
        self.assertEqual(data["skipped_duplicates"], 1)

    def test_preview_account_number_mismatch_returns_400(self):
        f = self._arion_csv(account_number="0370-13-999999")  # wrong account
        resp = self.client.post("/Import/preview", data={
            "user_id": self.user.id,
            "bank_account_id": self.bank_account.id,
            "bank": "arion",
            "file": f,
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("öðrum bankareikningi", resp.json()["detail"])

    def test_preview_unknown_bank_returns_400(self):
        f = self._arion_csv()
        resp = self.client.post("/Import/preview", data={
            "user_id": self.user.id,
            "bank_account_id": self.bank_account.id,
            "bank": "unknown_bank",
            "file": f,
        })
        self.assertEqual(resp.status_code, 400)

    def test_preview_wrong_extension_returns_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        f = SimpleUploadedFile("statement.pdf", b"not a spreadsheet", content_type="application/pdf")
        resp = self.client.post("/Import/preview", data={
            "user_id": self.user.id,
            "bank_account_id": self.bank_account.id,
            "bank": "arion",
            "file": f,
        })
        self.assertEqual(resp.status_code, 400)

    def test_preview_wrong_bank_account_returns_403(self):
        from associations.models import BankAccount
        other_bank_account = BankAccount.objects.create(
            association=self.other_association,
            name="Önnur", account_number="0370-13-000000",
            asset_account=self.asset_key,
        )
        f = self._arion_csv(account_number="0370-13-000000")
        resp = self.client.post("/Import/preview", data={
            "user_id": self.user.id,
            "bank_account_id": other_bank_account.id,
            "bank": "arion",
            "file": f,
        })
        self.assertEqual(resp.status_code, 403)

    def test_confirm_bulk_creates_transactions(self):
        from associations.models import Transaction
        rows = [
            {"date": "2026-03-15", "amount": "-245000.00", "description": "HS Veitur hf.", "reference": "280226"},
            {"date": "2026-03-10", "amount": "-180000.00", "description": "VÍS tryggingar", "reference": "290226"},
        ]
        resp = self.client.post(
            "/Import/confirm",
            data=json.dumps({
                "user_id": self.user.id,
                "bank_account_id": self.bank_account.id,
                "rows": rows,
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()["created"], 2)
        self.assertEqual(Transaction.objects.filter(bank_account=self.bank_account).count(), 2)
        tx = Transaction.objects.get(description="HS Veitur hf.")
        from decimal import Decimal
        self.assertEqual(tx.amount, Decimal("-245000.00"))
        self.assertEqual(tx.status, "IMPORTED")

    def test_confirm_wrong_bank_account_returns_403(self):
        from associations.models import BankAccount
        other_ba = BankAccount.objects.create(
            association=self.other_association,
            name="Önnur", account_number="0370-13-000000",
            asset_account=self.asset_key,
        )
        resp = self.client.post(
            "/Import/confirm",
            data=json.dumps({
                "user_id": self.user.id,
                "bank_account_id": other_ba.id,
                "rows": [],
            }),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)


class CategoryRuleModelTest(TestCase):
    def setUp(self):
        self.association = Association.objects.create(
            ssn="1234567890", name="Test Félag",
            address="Testgata 1", postal_code="101", city="Reykjavík"
        )
        self.category = Category.objects.create(name="Hitaveita", type="SHARED")

    def test_association_rule_created(self):
        from .models import CategoryRule
        rule = CategoryRule.objects.create(
            keyword="HS Veitur",
            category=self.category,
            association=self.association,
        )
        self.assertEqual(rule.keyword, "HS Veitur")
        self.assertFalse(rule.deleted)
        self.assertEqual(rule.association, self.association)

    def test_global_rule_has_null_association(self):
        from .models import CategoryRule
        rule = CategoryRule.objects.create(
            keyword="Orka",
            category=self.category,
            association=None,
        )
        self.assertIsNone(rule.association)


class CategoriserTest(TestCase):
    def setUp(self):
        self.association = Association.objects.create(
            ssn="9876543210", name="Félag B",
            address="Brautargata 2", postal_code="200", city="Kópavogur"
        )
        self.assoc2 = Association.objects.create(
            ssn="1111111119", name="Félag C",
            address="Vesturgata 3", postal_code="300", city="Akureyri"
        )
        self.cat_heat = Category.objects.create(name="Hitaveita", type="SHARED")
        self.cat_elec = Category.objects.create(name="Rafmagn", type="SHARED")
        self.cat_maint = Category.objects.create(name="Viðhald", type="SHARED")

    def test_normalise_vendor_strips_trailing_date(self):
        from .categoriser import normalise_vendor
        self.assertEqual(normalise_vendor("HS Veitur hf. 280226"), "hs veitur hf")

    def test_normalise_vendor_strips_reference_numbers(self):
        from .categoriser import normalise_vendor
        self.assertEqual(normalise_vendor("Orka náttúrunnar 12345678"), "orka náttúrunnar")

    def test_normalise_vendor_lowercases(self):
        from .categoriser import normalise_vendor
        self.assertEqual(normalise_vendor("VÍS Tryggingar"), "vís tryggingar")

    def test_categorise_row_association_rule_wins_over_global(self):
        from .models import CategoryRule
        from .categoriser import categorise_row
        CategoryRule.objects.create(keyword="Orka", category=self.cat_heat, association=self.association)
        CategoryRule.objects.create(keyword="Orka", category=self.cat_elec, association=None)
        # assoc rule first in list — simulating build_categorisation_context order
        rules = list(CategoryRule.objects.filter(deleted=False).order_by(
            django_models.Case(
                django_models.When(association=self.association, then=0),
                default=1,
            )
        ))
        result = categorise_row("Orka náttúrunnar", rules, {})
        self.assertEqual(result, self.cat_heat)

    def test_categorise_row_falls_back_to_history(self):
        from .categoriser import categorise_row
        history = {"orka náttúrunnar": self.cat_elec}
        result = categorise_row("Orka náttúrunnar 20260315", [], history)
        self.assertEqual(result, self.cat_elec)

    def test_categorise_row_returns_none_when_no_match(self):
        from .categoriser import categorise_row
        result = categorise_row("Óþekkt greiðsla", [], {})
        self.assertIsNone(result)

    def test_build_categorisation_context_returns_assoc_rules_first(self):
        from .models import CategoryRule
        from .categoriser import build_categorisation_context
        CategoryRule.objects.create(keyword="Global", category=self.cat_elec, association=None)
        CategoryRule.objects.create(keyword="Local", category=self.cat_heat, association=self.association)
        rules, history = build_categorisation_context(self.association)
        self.assertEqual(rules[0].association, self.association)
        self.assertIsNone(rules[1].association)

    def test_build_categorisation_context_excludes_deleted(self):
        from .models import CategoryRule
        from .categoriser import build_categorisation_context
        CategoryRule.objects.create(keyword="Dead", category=self.cat_elec, association=None, deleted=True)
        rules, history = build_categorisation_context(self.association)
        self.assertFalse(any(r.keyword == "Dead" for r in rules))

    def test_build_categorisation_context_history_from_categorised_transactions(self):
        from .models import CategoryRule, BankAccount, Transaction, TransactionStatus
        from .categoriser import build_categorisation_context, normalise_vendor
        bank = BankAccount.objects.create(
            association=self.association, name="Sparnaður", account_number="0111-26-123456"
        )
        Transaction.objects.create(
            bank_account=bank,
            date="2026-01-15",
            amount="-5000",
            description="HS Veitur hf. 280226",
            status=TransactionStatus.CATEGORISED,
            category=self.cat_heat,
        )
        _, history = build_categorisation_context(self.association)
        self.assertEqual(history.get(normalise_vendor("HS Veitur hf. 280226")), self.cat_heat)

    def test_build_categorisation_context_history_excludes_other_association(self):
        from .models import CategoryRule, BankAccount, Transaction, TransactionStatus
        from .categoriser import build_categorisation_context, normalise_vendor
        bank2 = BankAccount.objects.create(
            association=self.assoc2, name="Annar reikningur", account_number="0222-26-999999"
        )
        Transaction.objects.create(
            bank_account=bank2,
            date="2026-01-10",
            amount="-1000",
            description="HS Veitur hf. 100226",
            status=TransactionStatus.CATEGORISED,
            category=self.cat_heat,
        )
        _, history = build_categorisation_context(self.association)
        self.assertNotIn(normalise_vendor("HS Veitur hf. 100226"), history)


class ImportConfirmCategorisationTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="1234567890", name="Tester")
        self.association = Association.objects.create(
            ssn="0101013210", name="Flokkunarfélag",
            address="Flokkunargata 1", postal_code="101", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        self.bank = BankAccount.objects.create(
            association=self.association, name="Aðalreikningur", account_number="0111-26-000001"
        )
        self.cat = Category.objects.create(name="Hitaveita", type="SHARED")

    def _confirm(self, rows):
        return self.client.post(
            "/Import/confirm",
            data=json.dumps({"user_id": self.user.id, "bank_account_id": self.bank.id, "rows": rows}),
            content_type="application/json",
        )

    def test_import_with_matching_rule_categorises_transaction(self):
        from .models import CategoryRule
        CategoryRule.objects.create(keyword="HS Veitur", category=self.cat, association=self.association)
        resp = self._confirm([{"date": "2026-03-01", "amount": "-5000", "description": "HS Veitur hf. 280226", "reference": ""}])
        self.assertEqual(resp.status_code, 201)
        txn = Transaction.objects.filter(bank_account=self.bank).first()
        self.assertEqual(txn.category, self.cat)
        self.assertEqual(txn.status, TransactionStatus.CATEGORISED)

    def test_import_with_no_rule_leaves_status_imported(self):
        resp = self._confirm([{"date": "2026-03-02", "amount": "-1000", "description": "Óþekkt greiðsla", "reference": ""}])
        self.assertEqual(resp.status_code, 201)
        txn = Transaction.objects.filter(bank_account=self.bank).first()
        self.assertIsNone(txn.category)
        self.assertEqual(txn.status, TransactionStatus.IMPORTED)

    def test_import_with_history_match_categorises_transaction(self):
        Transaction.objects.create(
            bank_account=self.bank,
            date="2026-01-01",
            amount="-5000",
            description="HS Veitur hf. 010126",
            status=TransactionStatus.CATEGORISED,
            category=self.cat,
        )
        resp = self._confirm([{"date": "2026-03-01", "amount": "-5000", "description": "HS Veitur hf. 280226", "reference": ""}])
        self.assertEqual(resp.status_code, 201)
        txn = Transaction.objects.filter(bank_account=self.bank, date="2026-03-01").first()
        self.assertEqual(txn.category, self.cat)
        self.assertEqual(txn.status, TransactionStatus.CATEGORISED)


class CategoryRuleViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="3333333339", name="Reglumaður")
        self.superadmin = User.objects.create(kennitala="9999999999", name="Admin", is_superadmin=True)
        self.association = Association.objects.create(
            ssn="2020202020", name="Reglurfélag",
            address="Reglugata 5", postal_code="105", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        AssociationAccess.objects.create(user=self.superadmin, association=self.association, active=True)
        self.other_assoc = Association.objects.create(
            ssn="5050505050", name="Annað félag",
            address="Annargata 9", postal_code="200", city="Kópavogur"
        )
        self.cat = Category.objects.create(name="Hitaveita", type="SHARED")

    def test_get_returns_association_and_global_rules(self):
        from .models import CategoryRule
        CategoryRule.objects.create(keyword="Local", category=self.cat, association=self.association)
        CategoryRule.objects.create(keyword="Global", category=self.cat, association=None)
        resp = self.client.get(f"/CategoryRule/{self.user.id}?as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data["association_rules"]), 1)
        self.assertEqual(data["association_rules"][0]["keyword"], "Local")
        self.assertEqual(len(data["global_rules"]), 1)
        self.assertEqual(data["global_rules"][0]["keyword"], "Global")

    def test_get_excludes_deleted_rules(self):
        from .models import CategoryRule
        CategoryRule.objects.create(keyword="Dead", category=self.cat, association=self.association, deleted=True)
        resp = self.client.get(f"/CategoryRule/{self.user.id}?as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["association_rules"], [])

    def test_post_creates_association_rule(self):
        resp = self.client.post(
            "/CategoryRule",
            data=json.dumps({"user_id": self.user.id, "keyword": "VÍS", "category_id": self.cat.id, "is_global": False}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(data["keyword"], "VÍS")
        self.assertFalse(data["is_global"])

    def test_post_global_rule_by_superadmin(self):
        resp = self.client.post(
            "/CategoryRule",
            data=json.dumps({"user_id": self.superadmin.id, "keyword": "Orka", "category_id": self.cat.id, "is_global": True}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertTrue(data["is_global"])

    def test_post_global_rule_by_non_superadmin_returns_403(self):
        resp = self.client.post(
            "/CategoryRule",
            data=json.dumps({"user_id": self.user.id, "keyword": "Orka", "category_id": self.cat.id, "is_global": True}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_post_unknown_category_returns_400(self):
        resp = self.client.post(
            "/CategoryRule",
            data=json.dumps({"user_id": self.user.id, "keyword": "Test", "category_id": 99999, "is_global": False}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)

    def test_put_updates_keyword(self):
        from .models import CategoryRule
        rule = CategoryRule.objects.create(keyword="Gamalt", category=self.cat, association=self.association)
        resp = self.client.put(
            f"/CategoryRule/update/{rule.id}",
            data=json.dumps({"user_id": self.user.id, "keyword": "Nýtt", "category_id": self.cat.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        rule.refresh_from_db()
        self.assertEqual(rule.keyword, "Nýtt")

    def test_delete_soft_deletes_rule(self):
        from .models import CategoryRule
        rule = CategoryRule.objects.create(keyword="Eyða", category=self.cat, association=self.association)
        resp = self.client.delete(
            f"/CategoryRule/delete/{rule.id}",
            data=json.dumps({"user_id": self.user.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        rule.refresh_from_db()
        self.assertTrue(rule.deleted)

    def test_update_rule_of_other_association_returns_403(self):
        from .models import CategoryRule
        other_rule = CategoryRule.objects.create(
            keyword="Annað", category=self.cat, association=self.other_assoc
        )
        resp = self.client.put(
            f"/CategoryRule/update/{other_rule.id}",
            data=json.dumps({"user_id": self.user.id, "keyword": "Hacked", "category_id": self.cat.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)

    def test_update_nonexistent_rule_returns_404(self):
        resp = self.client.put(
            "/CategoryRule/update/99999",
            data=json.dumps({"user_id": self.user.id, "keyword": "X", "category_id": self.cat.id}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 404)


import datetime as _datetime_module


class ReportViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create(kennitala="7777777779", name="Skýrslumaður")
        self.association = Association.objects.create(
            ssn="7070707070", name="Skýrslufélag",
            address="Skýrslugata 1", postal_code="107", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        self.cat_heat = Category.objects.create(name="Hitaveita", type="SHARED")
        self.cat_elec = Category.objects.create(name="Rafmagn", type="SHARED")
        self.bank = BankAccount.objects.create(
            association=self.association,
            name="Aðalreikningur",
            account_number="0101-26-123456",
        )

    def _tx(self, amount, cat=None, date=None):
        from decimal import Decimal
        return Transaction.objects.create(
            bank_account=self.bank,
            date=date or _datetime_module.date(2026, 3, 15),
            amount=Decimal(str(amount)),
            description="Test",
            reference='',
            category=cat,
            status=TransactionStatus.CATEGORISED if cat else TransactionStatus.IMPORTED,
        )

    def test_income_and_expense_totals(self):
        from decimal import Decimal
        self._tx(400000)                         # income uncategorised
        self._tx(-95000, cat=self.cat_heat)      # expense categorised
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(Decimal(data["income_uncategorised"]), Decimal("400000"))
        self.assertEqual(len(data["expenses"]), 1)
        self.assertEqual(Decimal(data["expenses"][0]["actual"]), Decimal("95000"))

    def test_budget_comparison(self):
        from decimal import Decimal
        from .models import Budget, BudgetItem
        budget = Budget.objects.create(
            association=self.association, year=2026, version=1, is_active=True
        )
        BudgetItem.objects.create(budget=budget, category=self.cat_heat, amount=Decimal("1200000"))
        self._tx(-950000, cat=self.cat_heat)
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        expense = data["expenses"][0]
        self.assertEqual(Decimal(expense["budgeted"]), Decimal("1200000"))
        self.assertEqual(Decimal(expense["actual"]), Decimal("950000"))

    def test_budget_item_with_no_transactions_returns_zero_actual(self):
        from decimal import Decimal
        from .models import Budget, BudgetItem
        budget = Budget.objects.create(
            association=self.association, year=2026, version=1, is_active=True
        )
        BudgetItem.objects.create(budget=budget, category=self.cat_elec, amount=Decimal("600000"))
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        expense = next(e for e in data["expenses"] if e["category_id"] == self.cat_elec.id)
        self.assertEqual(Decimal(expense["actual"]), Decimal("0"))

    def test_expense_with_no_budget_returns_zero_budgeted(self):
        from decimal import Decimal
        self._tx(-85000, cat=self.cat_heat)
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(Decimal(data["expenses"][0]["budgeted"]), Decimal("0"))

    def test_uncategorised_income_and_expense(self):
        from decimal import Decimal
        self._tx(100000)
        self._tx(-50000)
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        data = resp.json()
        self.assertEqual(Decimal(data["income_uncategorised"]), Decimal("100000"))
        self.assertEqual(Decimal(data["expenses_uncategorised"]), Decimal("50000"))

    def test_year_param(self):
        self._tx(-100000, cat=self.cat_heat, date=_datetime_module.date(2025, 6, 1))
        resp = self.client.get(f"/Report/{self.user.id}?year=2025&as={self.association.id}")
        data = resp.json()
        self.assertEqual(data["year"], 2025)
        self.assertEqual(len(data["expenses"]), 1)

    def test_month_param_filters_to_single_month(self):
        from decimal import Decimal
        self._tx(-100000, cat=self.cat_heat, date=_datetime_module.date(2026, 3, 15))
        self._tx(-200000, cat=self.cat_heat, date=_datetime_module.date(2026, 4, 10))
        resp = self.client.get(
            f"/Report/{self.user.id}?year=2026&month=3&as={self.association.id}"
        )
        data = resp.json()
        self.assertEqual(Decimal(data["expenses"][0]["actual"]), Decimal("100000"))
        self.assertEqual(data["monthly"], [])

    def test_no_transactions_returns_zeros(self):
        from decimal import Decimal
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["income"], [])
        self.assertEqual(Decimal(data["income_uncategorised"]), Decimal("0"))
        self.assertEqual(data["expenses"], [])
        self.assertEqual(Decimal(data["expenses_uncategorised"]), Decimal("0"))
        self.assertEqual(len(data["monthly"]), 12)

    def test_monthly_breakdown(self):
        from decimal import Decimal
        self._tx(400000, date=_datetime_module.date(2026, 1, 10))
        self._tx(-95000, cat=self.cat_heat, date=_datetime_module.date(2026, 1, 15))
        resp = self.client.get(f"/Report/{self.user.id}?year=2026&as={self.association.id}")
        data = resp.json()
        jan = data["monthly"][0]
        self.assertEqual(jan["month"], 1)
        self.assertEqual(Decimal(jan["income"]), Decimal("400000"))
        self.assertEqual(Decimal(jan["expenses"]), Decimal("95000"))

    def test_superadmin_as_param(self):
        from decimal import Decimal
        superadmin = User.objects.create(
            kennitala="9999999998", name="Admin2", is_superadmin=True
        )
        self._tx(100000)  # income transaction in this association
        resp = self.client.get(
            f"/Report/{superadmin.id}?year=2026&as={self.association.id}"
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(Decimal(data["income_uncategorised"]), Decimal("100000"))


class ImporterKennitalaTest(TestCase):
    """Verify each parser extracts payer_kennitala from the correct column."""

    def _make_xlsx(self, rows):
        """Helper: build an in-memory xlsx with the given list-of-lists."""
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_arion_extracts_kennitala(self):
        from .importers import parse_arion
        rows = [
            ["Arion banki"],
            ["0133-26-111111"],
            [],
            ["Dagsetning", "Upphæð", "Texti", "Seðilnúmer", "Kennitala viðtakanda eða greiðanda"],
            ["1. jan. 2026", "45.000", "Húsgjöld", "12345", "1234567890"],
        ]
        result = parse_arion(self._make_xlsx(rows), "xlsx")
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["payer_kennitala"], "1234567890")

    def test_landsbankinn_extracts_kennitala(self):
        from .importers import parse_landsbankinn
        rows = [
            ["Landsbankinn"],
            ["Færslur á reikningi 0133-26-019111 ..."],
            ["2026"],
            [],
            ["Dags", "Upphæð", "Texti", "Tnr/Seðilnr.", "Kennitala"],
            ["01.01.2026", "45.000", "Húsgjöld", "99", "0987654321"],
        ]
        result = parse_landsbankinn(self._make_xlsx(rows), "xlsx")
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["payer_kennitala"], "0987654321")

    def test_islandsbanki_old_extracts_kennitala(self):
        from .importers import parse_islandsbanki
        rows = [
            [],
            [],
            [],
            [],
            ["Dags.", "Upph.ISK", "Mótaðili", "Tilvísun", "Kennitala móttakanda"],
            ["01.01.2026", "45.000", "Húsgjöld", "REF1", "5555555555"],
        ]
        result = parse_islandsbanki(self._make_xlsx(rows), "xlsx")
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["payer_kennitala"], "5555555555")

    def test_islandsbanki_new_returns_empty_kennitala(self):
        from .importers import parse_islandsbanki
        # New format: A4 == "Reikningsnúmer", account in B4, headers on row 12
        rows = [
            [],
            [],
            [],
            ["Reikningsnúmer", "0101-26-123456"],
            [], [], [], [], [], [], [],
            ["Dagsetning", "Upphæð", "Mótaðili", "Tilvísun"],
            ["01.01.2026", "45.000", "Húsgjöld", "REF1"],
        ]
        result = parse_islandsbanki(self._make_xlsx(rows), "xlsx")
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["payer_kennitala"], "")


class AutoMatchTest(TestCase):
    def setUp(self):
        from decimal import Decimal
        self.client = Client()
        self.user = User.objects.create(kennitala="1111111119", name="Admin")
        self.payer_user = User.objects.create(kennitala="1234567890", name="Jón Jónsson")
        self.association = Association.objects.create(
            ssn="2020202020", name="Próffélag",
            address="Prófgata 1", postal_code="101", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        self.bank = BankAccount.objects.create(
            association=self.association, name="Aðalreikningur", account_number="0101-26-000001"
        )
        self.budget = Budget.objects.create(
            association=self.association, year=2026, version=1, is_active=True
        )
        apt = Apartment.objects.create(
            association=self.association, anr="0101",
            share=Decimal("100"), share_2=Decimal("0"),
            share_3=Decimal("0"), share_eq=Decimal("100"),
        )
        ApartmentOwnership.objects.create(
            apartment=apt, user=self.payer_user,
            share=Decimal("100"), is_payer=True, deleted=False,
        )
        self.collection = Collection.objects.create(
            budget=self.budget, apartment=apt, payer=self.payer_user,
            month=3, amount_shared=Decimal("0"), amount_equal=Decimal("45000"),
            amount_total=Decimal("45000"), status=CollectionStatus.PENDING,
        )

    def _import_rows(self, rows):
        """POST /Import/confirm with given rows for self.bank."""
        return self.client.post(
            "/Import/confirm",
            data=json.dumps({
                "user_id": self.user.id,
                "bank_account_id": self.bank.id,
                "rows": rows,
            }),
            content_type="application/json",
        )

    def test_matching_kennitala_auto_matches(self):
        from decimal import Decimal
        resp = self._import_rows([{
            "date": "2026-03-10",
            "amount": "45000",
            "description": "Húsgjöld",
            "reference": "",
            "payer_kennitala": "1234567890",
        }])
        self.assertEqual(resp.status_code, 201)
        tx = Transaction.objects.get(bank_account=self.bank)
        self.assertEqual(tx.payer_kennitala, "1234567890")
        self.assertEqual(tx.status, TransactionStatus.RECONCILED)
        self.collection.refresh_from_db()
        self.assertEqual(self.collection.status, CollectionStatus.PAID)
        self.assertEqual(self.collection.paid_transaction_id, tx.id)

    def test_unknown_kennitala_stays_unmatched(self):
        self._import_rows([{
            "date": "2026-03-10", "amount": "45000",
            "description": "Húsgjöld", "reference": "",
            "payer_kennitala": "9999999999",
        }])
        tx = Transaction.objects.get(bank_account=self.bank)
        self.assertNotEqual(tx.status, TransactionStatus.RECONCILED)
        self.collection.refresh_from_db()
        self.assertEqual(self.collection.status, CollectionStatus.PENDING)

    def test_negative_amount_ignored(self):
        self._import_rows([{
            "date": "2026-03-10", "amount": "-45000",
            "description": "Útgreiðsla", "reference": "",
            "payer_kennitala": "1234567890",
        }])
        self.collection.refresh_from_db()
        self.assertEqual(self.collection.status, CollectionStatus.PENDING)

    def test_empty_kennitala_stays_unmatched(self):
        self._import_rows([{
            "date": "2026-03-10", "amount": "45000",
            "description": "Húsgjöld", "reference": "",
            "payer_kennitala": "",
        }])
        self.collection.refresh_from_db()
        self.assertEqual(self.collection.status, CollectionStatus.PENDING)


class CollectionGenerateViewTest(TestCase):
    def setUp(self):
        from decimal import Decimal
        self.client = Client()
        self.user = User.objects.create(kennitala="3333333339", name="Admin")
        self.payer1 = User.objects.create(kennitala="1111111111", name="Greiðandi 1")
        self.payer2 = User.objects.create(kennitala="2222222222", name="Greiðandi 2")
        self.association = Association.objects.create(
            ssn="3030303030", name="Kynningarfélag",
            address="Kynningargata 1", postal_code="101", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        self.budget = Budget.objects.create(
            association=self.association, year=2026, version=1, is_active=True
        )
        cat = Category.objects.create(name="Sameiginlegt", type="EQUAL")
        BudgetItem.objects.create(budget=self.budget, category=cat, amount=Decimal("600000"))
        self.apt1 = Apartment.objects.create(
            association=self.association, anr="0101", fnr="F000001",
            share=Decimal("0"), share_2=Decimal("0"),
            share_3=Decimal("0"), share_eq=Decimal("50"),
        )
        self.apt2 = Apartment.objects.create(
            association=self.association, anr="0102", fnr="F000002",
            share=Decimal("0"), share_2=Decimal("0"),
            share_3=Decimal("0"), share_eq=Decimal("50"),
        )
        ApartmentOwnership.objects.create(
            apartment=self.apt1, user=self.payer1, share=Decimal("100"), is_payer=True, deleted=False
        )
        ApartmentOwnership.objects.create(
            apartment=self.apt2, user=self.payer2, share=Decimal("100"), is_payer=True, deleted=False
        )

    def _post(self, month=3, year=2026):
        return self.client.post(
            "/Collection/generate",
            data=json.dumps({"user_id": self.user.id, "month": month, "year": year}),
            content_type="application/json",
        )

    def test_generates_correct_number_of_items(self):
        resp = self._post()
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(data["created"], 2)
        self.assertEqual(data["skipped"], 0)
        self.assertEqual(Collection.objects.filter(budget=self.budget, month=3).count(), 2)

    def test_is_idempotent(self):
        self._post()
        resp = self._post()
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(data["skipped"], 2)
        self.assertEqual(Collection.objects.filter(budget=self.budget, month=3).count(), 2)

    def test_404_no_active_budget(self):
        resp = self.client.post(
            "/Collection/generate",
            data=json.dumps({"user_id": self.user.id, "month": 3, "year": 2025}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 404)

    def test_payer_from_current_is_payer_ownership(self):
        from decimal import Decimal
        self._post()
        col = Collection.objects.get(budget=self.budget, apartment=self.apt1, month=3)
        self.assertEqual(col.payer, self.payer1)

    def test_amounts_calculated_correctly(self):
        from decimal import Decimal
        self._post()
        col = Collection.objects.get(budget=self.budget, apartment=self.apt1, month=3)
        # Budget 600000 EQUAL, apt1 share_eq=50 → 600000 * 50 / 100 = 300000
        self.assertEqual(col.amount_equal, Decimal("300000"))
        self.assertEqual(col.amount_total, Decimal("300000"))


class CollectionMatchViewTest(TestCase):
    def setUp(self):
        from decimal import Decimal
        import datetime as dt
        self.client = Client()
        self.user = User.objects.create(kennitala="4444444449", name="Admin")
        self.payer = User.objects.create(kennitala="5555555555", name="Greiðandi")
        self.association = Association.objects.create(
            ssn="4040404040", name="Samræmisfélag",
            address="Samræmisgata 1", postal_code="101", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        self.bank = BankAccount.objects.create(
            association=self.association, name="Aðalreikningur", account_number="0101-26-000004"
        )
        self.budget = Budget.objects.create(
            association=self.association, year=2026, version=1, is_active=True
        )
        apt = Apartment.objects.create(
            association=self.association, anr="0101", fnr="F000003",
            share=Decimal("0"), share_2=Decimal("0"),
            share_3=Decimal("0"), share_eq=Decimal("100"),
        )
        ApartmentOwnership.objects.create(
            apartment=apt, user=self.payer, share=Decimal("100"), is_payer=True, deleted=False
        )
        self.collection = Collection.objects.create(
            budget=self.budget, apartment=apt, payer=self.payer,
            month=3, amount_shared=Decimal("0"), amount_equal=Decimal("45000"),
            amount_total=Decimal("45000"), status=CollectionStatus.PENDING,
        )
        self.tx = Transaction.objects.create(
            bank_account=self.bank,
            date=dt.date(2026, 3, 10),
            amount=Decimal("45000"),
            description="Húsgjöld",
            reference="",
            status=TransactionStatus.IMPORTED,
        )

    def _post(self, collection_id=None, transaction_id=None):
        return self.client.post(
            "/Collection/match",
            data=json.dumps({
                "user_id": self.user.id,
                "collection_id": collection_id or self.collection.id,
                "transaction_id": transaction_id or self.tx.id,
            }),
            content_type="application/json",
        )

    def test_manual_match_sets_paid_and_reconciled(self):
        resp = self._post()
        self.assertEqual(resp.status_code, 200)
        self.collection.refresh_from_db()
        self.tx.refresh_from_db()
        self.assertEqual(self.collection.status, CollectionStatus.PAID)
        self.assertEqual(self.collection.paid_transaction_id, self.tx.id)
        self.assertEqual(self.tx.status, TransactionStatus.RECONCILED)

    def test_returns_400_if_collection_already_paid(self):
        self.collection.status = CollectionStatus.PAID
        self.collection.paid_transaction = self.tx
        self.collection.save()
        resp = self._post()
        self.assertEqual(resp.status_code, 400)

    def test_returns_400_if_transaction_not_positive(self):
        from decimal import Decimal
        self.tx.amount = Decimal("-100")
        self.tx.save()
        resp = self._post()
        self.assertEqual(resp.status_code, 400)


class CollectionViewMonthTest(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        self.client = Client()
        self.user = User.objects.create(kennitala="6666666669", name="Admin")
        self.payer = User.objects.create(kennitala="7777777771", name="Greiðandi")
        self.association = Association.objects.create(
            ssn="5050505050", name="Mánuðarfélag",
            address="Mánuðargata 1", postal_code="101", city="Reykjavík"
        )
        AssociationAccess.objects.create(user=self.user, association=self.association, active=True)
        self.bank = BankAccount.objects.create(
            association=self.association, name="Aðalreikningur", account_number="0101-26-000006"
        )
        self.budget = Budget.objects.create(
            association=self.association, year=2026, version=1, is_active=True
        )
        apt = Apartment.objects.create(
            association=self.association, anr="0101", fnr="F000011",
            share=Decimal("0"), share_2=Decimal("0"),
            share_3=Decimal("0"), share_eq=Decimal("100"),
        )
        ApartmentOwnership.objects.create(
            apartment=apt, user=self.payer,
            share=Decimal("100"), is_payer=True, deleted=False
        )
        self.collection = Collection.objects.create(
            budget=self.budget, apartment=apt, payer=self.payer,
            month=3, amount_shared=Decimal("0"), amount_equal=Decimal("45000"),
            amount_total=Decimal("45000"), status=CollectionStatus.PENDING,
        )
        # Unmatched income transaction — positive, not RECONCILED, not linked
        self.unmatched_tx = Transaction.objects.create(
            bank_account=self.bank,
            date=dt.date(2026, 3, 12),
            amount=Decimal("41000"),
            description="Húsgjöld mars - Sigríður",
            reference="",
            status=TransactionStatus.IMPORTED,
        )
        # Expense transaction — should NOT appear in unmatched
        Transaction.objects.create(
            bank_account=self.bank,
            date=dt.date(2026, 3, 5),
            amount=Decimal("-10000"),
            description="Raf",
            reference="",
            status=TransactionStatus.CATEGORISED,
        )

    def test_month_mode_returns_collection_rows(self):
        resp = self.client.get(
            f"/Collection/{self.user.id}?month=3&year=2026&as={self.association.id}"
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["month"], 3)
        self.assertEqual(data["year"], 2026)
        self.assertEqual(len(data["rows"]), 1)
        row = data["rows"][0]
        self.assertEqual(row["collection_id"], self.collection.id)
        self.assertEqual(row["status"], "PENDING")
        self.assertIsNone(row["paid_transaction_id"])

    def test_month_mode_returns_unmatched_transactions(self):
        resp = self.client.get(
            f"/Collection/{self.user.id}?month=3&year=2026&as={self.association.id}"
        )
        data = resp.json()
        self.assertEqual(len(data["unmatched"]), 1)
        self.assertEqual(data["unmatched"][0]["transaction_id"], self.unmatched_tx.id)

    def test_month_mode_excludes_expenses_from_unmatched(self):
        resp = self.client.get(
            f"/Collection/{self.user.id}?month=3&year=2026&as={self.association.id}"
        )
        data = resp.json()
        amounts = [float(u["amount"]) for u in data["unmatched"]]
        self.assertNotIn(-10000.0, amounts)

    def test_summary_mode_still_returns_computed_rows(self):
        resp = self.client.get(
            f"/Collection/{self.user.id}?summary=1&as={self.association.id}"
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertIn("rows", data)
        self.assertIn("budget_summary", data)
